# -*- coding: utf-8 -*-
"""Smoke test de la refonte profonde (2026-06-10, 2e passe).

Vérifie SANS réseau et SANS démarrer les workers :
  1. Chien de garde : diagnostic des robots (panne / muet / volontairement
     éteint / en retrait derrière le serveur).
  2. Présence serveur : arbitrage desktop/serveur (battement de cœur).
  3. Docteur DNS : analyse SPF/DKIM/DMARC/MX (verdicts purs).
  4. Performance par modèle : agrégation envois × réponses.
  5. Export Excel commun : fichier généré, lignes comptées, surlignage.
  6. Clés API : plus AUCUNE clé en dur dans le code.
  7. Nouveaux endpoints exposés (mail_dns_check, funnel_by_template).

Usage :  python scripts/smoke_refonte_v2.py
"""
from __future__ import annotations

import sys
import tempfile
from datetime import datetime, timedelta, timezone
from pathlib import Path

HERE = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(HERE))
CORE = HERE.parent / "triskell-core"
if CORE.exists():
    sys.path.insert(0, str(CORE))

PASS, FAIL = [], []


def check(label: str, cond: bool, detail: str = "") -> None:
    (PASS if cond else FAIL).append(label)
    print(f"  {'OK  ' if cond else 'FAIL'}- {label} {detail if not cond else ''}")


print("1) Chien de garde…")
from triskell_command.integrations.worker_watchdog import diagnose_workers  # noqa: E402

_now = datetime.now(timezone.utc)
_old = (_now - timedelta(hours=5)).isoformat(timespec="seconds")
_recent = (_now - timedelta(minutes=10)).isoformat(timespec="seconds")
workers = [
    {"name": "ok_one", "label": "Robot sain", "running": True,
     "health": "healthy", "last_run_at": _recent, "last_run_result": {}},
    {"name": "warn_one", "label": "Robot en panne", "running": True,
     "health": "warning", "last_run_at": _recent,
     "last_run_result": {"error": "JWT expired"}},
    {"name": "mute_one", "label": "Robot muet", "running": True,
     "health": "healthy", "last_run_at": _old, "last_run_result": {}},
    {"name": "stopped_one", "label": "Robot arrêté", "running": False,
     "health": "healthy", "last_run_at": _recent, "last_run_result": {}},
    {"name": "autopilot_runner", "label": "Auto-pilote", "running": False,
     "health": "healthy", "last_run_at": "",
     "last_run_result": {"skipped_reason": "disabled"}},
    {"name": "desktop_defer", "label": "Robot en retrait", "running": True,
     "health": "healthy", "last_run_at": _old,
     "last_run_result": {"skipped_reason": "server_active"}},
]
problems = diagnose_workers(workers)
names = {p["name"] for p in problems}
check("panne détectée (warning)", "warn_one" in names)
check("robot muet détecté (5h sans passage)", "mute_one" in names)
check("robot arrêté détecté", "stopped_one" in names)
check("robot sain non alerté", "ok_one" not in names)
check("auto-pilote désactivé = pas une panne", "autopilot_runner" not in names)
check("robot en retrait (serveur actif) = pas une panne",
      "desktop_defer" not in names)

# Seuil de silence PAR robot (les robots distants tolèrent plus que 3 h)
_h10 = (_now - timedelta(hours=10)).isoformat(timespec="seconds")
slow_workers = [
    {"name": "remote_ok", "label": "Distant 5h", "running": True,
     "health": "healthy", "last_run_at": _old, "last_run_result": {},
     "stale_after_hours": 9},
    {"name": "remote_mute", "label": "Distant 10h", "running": True,
     "health": "healthy", "last_run_at": _h10, "last_run_result": {},
     "stale_after_hours": 9},
]
slow_names = {p["name"] for p in diagnose_workers(slow_workers)}
check("seuil par robot : 5h de silence tolérées (stale_after_hours=9)",
      "remote_ok" not in slow_names)
check("seuil par robot : 10h de silence = muet (stale_after_hours=9)",
      "remote_mute" in slow_names)

print("1bis) Battement de cœur du Phare…")
from triskell_command.integrations.phare import heartbeat as ph_hb  # noqa: E402

_tick_fresh = (_now - timedelta(hours=2)).isoformat(timespec="seconds")
_tick_old = (_now - timedelta(hours=12)).isoformat(timespec="seconds")
w = ph_hb.evaluate({"enabled": True, "last_tick_at": _tick_fresh}, now=_now)
check("tick frais (2h) → robot Phare sain",
      w is not None and w["health"] == "healthy" and w["running"])
w = ph_hb.evaluate({"enabled": True, "last_tick_at": _tick_old}, now=_now)
check("tick vieux (12h) → robot Phare en panne",
      w is not None and w["health"] == "error" and not w["running"])
check("…avec une erreur qui pointe vers GitHub Actions",
      w is not None and "github" in (w["last_run_result"].get("error") or "").lower())
w = ph_hb.evaluate({"enabled": False, "last_tick_at": _tick_old}, now=_now)
check("Phare désactivé → visible mais jamais alerté",
      w is not None and (w["last_run_result"].get("skipped_reason") == "disabled"))
_today = _now.date().isoformat()
w = ph_hb.evaluate({"enabled": True,
                    "scheduler_log": {"algo_watch:": _today}}, now=_now)
check("pas de battement mais mission globale aujourd'hui → sain (fallback)",
      w is not None and w["health"] == "healthy")
_j3 = (_now - timedelta(days=3)).date().isoformat()
w = ph_hb.evaluate({"enabled": True,
                    "scheduler_log": {"algo_watch:": _j3}}, now=_now)
check("dernière mission globale il y a 3 jours → panne (fallback)",
      w is not None and w["health"] == "error")
check("config vide → pas de robot virtuel (pas de fausse alerte)",
      ph_hb.evaluate({}, now=_now) is None)
# Le pseudo-robot doit passer le diagnostic du chien de garde sans bruit
w_ok = ph_hb.evaluate({"enabled": True, "last_tick_at": _old}, now=_now)
check("Phare : 5h sans tick GitHub = toléré par le chien de garde",
      w_ok is not None and not diagnose_workers([w_ok]))
w_ko = ph_hb.evaluate({"enabled": True, "last_tick_at": _tick_old}, now=_now)
check("Phare : 12h sans tick = alerte du chien de garde",
      w_ko is not None and len(diagnose_workers([w_ko])) == 1)

print("1ter) Construction auto Pixel Pros (simulation à blanc)…")
from triskell_command.integrations.pixelpros import auto_builder as ab  # noqa: E402
from triskell_command.integrations.pixelpros import repo as pp_repo  # noqa: E402


class _FakeSb:
    """Faux client base : réglages en mémoire, aucune requête réseau."""
    def __init__(self, settings): self._s = dict(settings)
    is_authenticated = True
    def get_shared_setting(self, key, default=None): return self._s.get(key, default)
    def set_shared_setting(self, key, value): self._s[key] = value


_paid_old = (_now - timedelta(hours=2)).isoformat(timespec="seconds")
_paid_fresh = (_now - timedelta(seconds=60)).isoformat(timespec="seconds")
fake_intakes = [
    {"id": "i_fresh", "status": "paid", "stripe_paid_at": _paid_fresh,
     "business_name": "Boulangerie Toute Fraîche"},
    {"id": "i_old", "status": "paid", "stripe_paid_at": _paid_old,
     "business_name": "Garage Dupont"},
    {"id": "i_tried", "status": "paid", "stripe_paid_at": _paid_old,
     "business_name": "Déjà Tenté SARL"},
    {"id": "i_broken", "status": "paid", "stripe_paid_at": _paid_old,
     "business_name": "Échec Co"},
]
dispatched, failed_marks, notifs = [], [], []
_fake_sb = _FakeSb({
    ab.SETTING_ENABLED: True,
    ab.SETTING_STATE: {"i_tried": _paid_old},
})
_saved_ab = (ab._get_client, pp_repo.list_intakes, pp_repo.dispatch_build,
             pp_repo.mark_failed, ab._notify)
try:
    ab._get_client = lambda: _fake_sb
    pp_repo.list_intakes = lambda **kw: list(fake_intakes)

    def _fake_dispatch(iid):
        dispatched.append(iid)
        if iid == "i_broken":
            return False, "builder introuvable (test)"
        return True, "lancé"
    pp_repo.dispatch_build = _fake_dispatch
    pp_repo.mark_failed = (
        lambda iid, **kw: failed_marks.append(iid) or True)
    ab._notify = lambda title, body, priority="normal": notifs.append(
        (title, priority))

    # OFF → skipped proprement, rien touché
    _fake_sb._s[ab.SETTING_ENABLED] = False
    r_off = ab.tick(now=_now)
    check("auto-build OFF → skipped 'disabled', zéro lancement",
          r_off.get("skipped_reason") == "disabled" and not dispatched)

    # ON → la vieille part, la fraîche patiente, la déjà-tentée est ignorée
    _fake_sb._s[ab.SETTING_ENABLED] = True
    r_on = ab.tick(now=_now)
    check("payé il y a 2h → construction lancée", "i_old" in r_on["launched"])
    check("payé il y a 1 min → délai de grâce respecté",
          "i_fresh" not in dispatched and r_on["waiting_grace"] >= 1)
    check("déjà tenté → JAMAIS relancé tout seul",
          "i_tried" not in dispatched and r_on["already_tried"] >= 1)
    check("échec de lancement → commande marquée 'failed'",
          failed_marks == ["i_broken"])
    check("échec → alerte haute priorité envoyée",
          any(p == "high" for (_t, p) in notifs))
    check("lancement → notification d'information envoyée",
          any(p == "low" for (_t, p) in notifs))
    check("tentatives persistées (i_old et i_broken marqués)",
          "i_old" in _fake_sb._s[ab.SETTING_STATE]
          and "i_broken" in _fake_sb._s[ab.SETTING_STATE])

    # Re-tick : plus rien à faire (anti-boucle)
    dispatched.clear()
    r_again = ab.tick(now=_now)
    check("re-passage → aucune relance (anti-boucle)",
          not dispatched and not r_again["launched"])
finally:
    (ab._get_client, pp_repo.list_intakes, pp_repo.dispatch_build,
     pp_repo.mark_failed, ab._notify) = _saved_ab

print("2) Présence serveur…")
import os  # noqa: E402
from triskell_command.integrations import server_presence as sp  # noqa: E402


class _FakeClient:
    def __init__(self, hb): self._hb = hb
    def get_shared_setting(self, key, default=None): return self._hb
    def set_shared_setting(self, key, value): self._hb = value


os.environ.pop("TRISKELL_IS_SERVER", None)
fresh_hb = {"at": _now.isoformat(timespec="seconds"), "host": "srv"}
stale_hb = {"at": (_now - timedelta(hours=1)).isoformat(timespec="seconds"),
            "host": "srv"}
check("desktop s'efface si serveur vivant",
      sp.should_defer_to_server(_FakeClient(fresh_hb)) is True)
check("desktop reprend si serveur mort (>15 min)",
      sp.should_defer_to_server(_FakeClient(stale_hb)) is False)
check("sans accès base → comportement historique",
      sp.should_defer_to_server(None) is False)
sp.mark_server_process()
check("le serveur lui-même ne s'efface jamais",
      sp.should_defer_to_server(_FakeClient(fresh_hb)) is False)
os.environ.pop("TRISKELL_IS_SERVER", None)

print("3) Docteur DNS…")
from triskell_command.integrations.mail_dns_doctor import analyze_records  # noqa: E402

good = analyze_records(
    "triskell-studio.fr",
    spf_txts=["v=spf1 include:_spf-eu.ionos.com ~all"],
    dmarc_txts=["v=DMARC1; p=quarantine; rua=mailto:contact@triskell-studio.fr"],
    dkim_found_selector="s1-ionos",
    has_mx=True,
)
check("domaine bien configuré → 4/4", good["score"] == "4/4" and good["all_good"])
bad = analyze_records("nu.fr", spf_txts=[], dmarc_txts=[],
                      dkim_found_selector=None, has_mx=False)
check("domaine nu → 0/4 avec conseils",
      bad["score"] == "0/4" and all(c["advice"] or c["ok"] is False
                                     for c in bad["checks"]))
weak = analyze_records("x.fr", spf_txts=["v=spf1 ?all"],
                       dmarc_txts=["v=DMARC1; p=none"],
                       dkim_found_selector="s1", has_mx=True)
spf_check = next(c for c in weak["checks"] if c["id"] == "spf")
dmarc_check = next(c for c in weak["checks"] if c["id"] == "dmarc")
check("SPF laxiste signalé", "?all" in spf_check["advice"])
check("DMARC p=none signalé", "p=none" in dmarc_check["advice"]
      or "quarantine" in dmarc_check["advice"])

print("4) Performance par modèle…")
from triskell_command.integrations.funnel_metrics import (  # noqa: E402
    aggregate_template_performance,
)

sent = [
    {"prospect_id": "p1", "ts": "2026-06-01T10:00:00", "template_key": "pixel_pros_v1"},
    {"prospect_id": "p2", "ts": "2026-06-01T11:00:00", "template_key": "pixel_pros_v1"},
    {"prospect_id": "p3", "ts": "2026-06-02T10:00:00", "template_key": "carnet_v2"},
    {"prospect_id": "p1", "ts": "2026-06-05T10:00:00", "template_key": "relance_j5",
     "extra": {}},
]
replies = [
    # p1 répond APRÈS la relance → créditée à relance_j5
    {"prospect_id": "p1", "ts": "2026-06-06T09:00:00",
     "extra": {"classification": {"category": "interested"}}},
    # p3 répond après carnet_v2
    {"prospect_id": "p3", "ts": "2026-06-03T09:00:00",
     "extra": {"classification": {"category": "no"}}},
]
rows = aggregate_template_performance(sent, replies)
by_tpl = {r["template"]: r for r in rows}
check("envois comptés par modèle",
      by_tpl["pixel_pros_v1"]["sent"] == 2 and by_tpl["carnet_v2"]["sent"] == 1)
check("réponse créditée au DERNIER modèle envoyé",
      by_tpl["relance_j5"]["replies"] == 1
      and by_tpl["pixel_pros_v1"]["replies"] == 0)
check("intéressé compté", by_tpl["relance_j5"]["interested"] == 1)
check("taux calculé", by_tpl["carnet_v2"]["reply_rate"] == 100.0)

print("5) Export Excel commun…")
from triskell_command.integrations.hunt_exports import write_xlsx  # noqa: E402

with tempfile.TemporaryDirectory() as td:
    out = Path(td) / "test.xlsx"
    n = write_xlsx(out, sheet_title="Test",
                   headers=["A", "B"], rows=[[1, "x"], [2, "y"], [3, "z"]],
                   widths=[10, 20], highlight=[False, True, False])
    check("fichier Excel généré", out.exists() and out.stat().st_size > 0)
    check("lignes comptées", n == 3)
    from openpyxl import load_workbook
    wb = load_workbook(out)
    ws = wb.active
    check("entête + données présentes",
          ws["A1"].value == "A" and ws["A4"].value == 3)
    check("ligne 2 surlignée (ambre)",
          ws["A3"].fill.fgColor.rgb in ("00FEF3C7", "FFFEF3C7"))

print("6) Plus de clés en dur…")
import re  # noqa: E402

leaks = []
for f in [HERE / "triskell_command/web/api.py",
          HERE / "triskell_command/integrations/chasseur_createurs.py",
          HERE / "triskell_command/integrations/prospecteur_google.py"]:
    src = f.read_text(encoding="utf-8")
    if re.search(r'"AIzaSy[A-Za-z0-9_\-]{20,}"|"sk-[a-f0-9]{28,}"', src):
        leaks.append(f.name)
check("aucune clé API en dur dans le code", not leaks, f"(fuites : {leaks})")

print("7) Nouveaux endpoints…")
from triskell_command.web.api import Api  # noqa: E402

api = Api()
for ep in ("mail_dns_check", "funnel_by_template",
           "chasseur_createurs_push_to_prospects",
           "prospecteur_google_push_to_prospects",
           "phare_automerge_get", "phare_automerge_set"):
    check(f"endpoint {ep} exposé", callable(getattr(api, ep, None)))
# Sans domaine : l'endpoint retombe sur le domaine d'envoi configuré
# (machine avec SMTP → vraie vérif), sinon il refuse avec un message clair.
r = api.mail_dns_check({"domain": ""})
check("mail_dns_check sans domaine → réponse propre (fallback ou refus clair)",
      (r.get("ok") is True and r.get("checks"))
      or (r.get("ok") is False and "domaine" in (r.get("error") or "").lower()))

print()
print(f"{len(PASS)} OK / {len(FAIL)} échec(s)")
sys.exit(1 if FAIL else 0)
