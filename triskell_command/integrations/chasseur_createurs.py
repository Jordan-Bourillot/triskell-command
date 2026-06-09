"""Chasseur Créateur — découvre des créateurs de contenu (YouTube / Instagram /
Facebook) dans une niche donnée et récupère leurs mails publics.

Adapté depuis le programme desktop `chasseur-createurs/main.py`. On garde
l'architecture "hunt en arrière-plan" du Chasseur PME (cf chasseur.py) :
chaque chasse tourne dans un thread daemon, persiste son état en JSON dans
`~/.triskell-command/chasseur_createurs/hunts/*.json`, et l'UI poll l'état
toutes les 2 secondes.

Plateformes :
  - YouTube  : API officielle (clé Data API v3). Le plus fiable, c'est le mode
               recommandé. Scrape la description chaîne + page /about + suit
               les liens externes (linktree, beacons, etc.) + descriptions
               des 3 dernières vidéos.
  - Instagram : via `instagrapi` (lib non-officielle). Nécessite un login.
                Fragile, peut être bloqué par Meta. Optionnel.
  - Facebook : via `facebook_scraper`. Fragile pour les mêmes raisons.
                Optionnel.

Pour rester légal en B2B FR : on contacte uniquement des mails publiés
volontairement par le créateur sur ses propres pages publiques.
"""
from __future__ import annotations

import csv
import json
import logging
import os
import random
import re
import threading
import time
import uuid
from dataclasses import asdict, dataclass, field
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Callable

import requests

logger = logging.getLogger(__name__)


CHASSEUR_C_DIR = Path.home() / ".triskell-command" / "chasseur_createurs"
HUNTS_DIR = CHASSEUR_C_DIR / "hunts"
EXPORTS_DIR = CHASSEUR_C_DIR / "exports"

# Clé YouTube par défaut (peut être surchargée via la variable d'env
# YOUTUBE_API_KEY ou dans le payload de la chasse). C'est la clé qui était
# déjà utilisée dans le programme desktop ; elle a un quota de 10 000 unités
# par jour, ce qui permet ~80 chasses moyennes.
DEFAULT_YOUTUBE_API_KEY = "AIzaSyCu8xcAyxf091ureRYpaqzMtIKda6TAg10"

EMAIL_REGEX = r"[\w\.\-+]+@[\w\.\-]+\.\w+"

# Domaines "bio link" classiques qu'on suit pour aller chercher des mails
# qui ne sont pas directement dans la description YouTube.
BIO_DOMAINS = [
    "linktr.ee", "linktree.com", "beacons.ai", "beacons.page", "about.me",
    "carrd.co", "lnk.bio", "bio.link", "msha.ke", "koji.to",
    "flow.page", "allmylinks.com", "campsite.bio", "many.link",
    "snipfeed.co", "stan.store", "potion.so",
]

# Mails qu'on jette à la poubelle (faux positifs habituels)
EMAIL_BLACKLIST_PATTERNS = [
    "@youtube", "@google.", "@gstatic", "noreply", "no-reply",
    "example.com", "@sentry.io", "@cloudfront", "@amazonaws",
    "@bootstrapcdn", "@jquery", "@w3.org",
]

UA_WEB = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/120.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "fr-FR,fr;q=0.9,en;q=0.8",
    "Cookie": "CONSENT=YES+",
}

# Codes ISO acceptés pour le filtre francophone. "ALL" = tous francophones
# (pas de regionCode côté API, on s'appuie alors sur relevanceLanguage='fr'
# + filtre langdetect côté Python).
FRANCOPHONE_COUNTRIES = {
    "FR", "BE", "CH", "CA", "LU", "MC",
    "MA", "DZ", "TN", "SN", "CI",
}


def _normalize_pays(pays: str | None) -> str:
    """Renvoie un code ISO valide en majuscules, ou 'FR' par défaut."""
    code = (pays or "FR").strip().upper()
    if code == "ALL":
        return "ALL"
    if code not in FRANCOPHONE_COUNTRIES:
        return "FR"
    return code


def _is_french_text(text: str) -> bool:
    """Détecte si un texte est en français. Si langdetect indisponible ou
    en erreur, renvoie True (on garde la chaîne pour ne pas écarter à tort).
    """
    sample = (text or "").strip()
    if not sample:
        return True
    try:
        from langdetect import detect, DetectorFactory
        DetectorFactory.seed = 0  # détection déterministe
        return detect(sample) == "fr"
    except ImportError:
        return True
    except Exception:
        return True


def ensure_dirs() -> None:
    HUNTS_DIR.mkdir(parents=True, exist_ok=True)
    EXPORTS_DIR.mkdir(parents=True, exist_ok=True)


# ---------------------------------------------------------------------------
# Modèles
# ---------------------------------------------------------------------------
@dataclass
class Creator:
    """Un créateur trouvé dans une chasse."""
    platform: str               # youtube / instagram / facebook
    handle: str                 # channel_id ou username
    name: str                   # nom affiché de la chaîne / compte
    subscribers: int = 0        # abonnés / followers / likes
    url: str = ""               # URL publique du compte
    email: str = ""             # mail principal trouvé
    emails_extra: list[str] = field(default_factory=list)
    description: str = ""       # bio / description courte
    country: str = ""
    video_count: int = 0
    created_at: str = ""        # date de création du compte (ISO)
    external_links: list[str] = field(default_factory=list)

    def to_dict(self) -> dict:
        return asdict(self)


@dataclass
class CreatorHunt:
    """Une session de chasse aux créateurs."""
    id: str
    label: str
    created_at: str
    status: str = "pending"     # pending / searching / enriching / done / error
    progress: int = 0           # 0-100
    log: list[str] = field(default_factory=list)
    filters: dict = field(default_factory=dict)
    stats: dict = field(default_factory=dict)
    creators: list[dict] = field(default_factory=list)
    error: str = ""

    @property
    def path(self) -> Path:
        return HUNTS_DIR / f"{self.id}.json"

    def save(self) -> None:
        ensure_dirs()
        self.path.write_text(
            json.dumps(asdict(self), ensure_ascii=False, indent=2),
            encoding="utf-8",
        )

    @classmethod
    def load(cls, hunt_id: str) -> "CreatorHunt | None":
        p = HUNTS_DIR / f"{hunt_id}.json"
        if not p.exists():
            return None
        try:
            from .hunt_zombies import reconcile_hunt_file
            data = reconcile_hunt_file(p, is_running=is_running(hunt_id))
            if not data:
                return None
            return cls(**data)
        except Exception as exc:
            logger.warning("creator hunt.load failed %s: %s", hunt_id, exc)
            return None


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def _clean_emails(emails: set[str]) -> list[str]:
    """Vire les mails techniques / faux positifs.

    Toutes les extractions du module (YouTube, Instagram, Facebook)
    convergent ici. On applique DEUX couches :
      1. le filtre central de triskell_core (`email_filter.clean_email`) —
         la source de vérité anti-fausses-adresses (fragments d'URL,
         domaines plateforme/factices, local-parts suspects, préfixe www.) ;
         c'est le même filtre qui protège Obélisk depuis le 2026-05-22.
      2. la blacklist locale historique, en complément.
    Si triskell_core est absent (dev sans le paquet), seule la couche 2
    s'applique — comme avant.
    """
    try:
        from triskell_core.prospect.enrichers.email_filter import clean_email
    except ImportError:
        clean_email = None  # fallback : blacklist locale seule
    out: list[str] = []
    for e in emails:
        e = (e or "").strip().lower()
        if not e:
            continue
        if clean_email is not None:
            e = clean_email(e) or ""
            if not e:
                continue
        if any(p in e for p in EMAIL_BLACKLIST_PATTERNS):
            continue
        if e in out:
            continue
        out.append(e)
    return out


def _get_youtube_api_key(payload_key: str | None = None) -> str:
    if payload_key and payload_key.strip():
        return payload_key.strip()
    env_key = (os.environ.get("YOUTUBE_API_KEY") or "").strip()
    if env_key:
        return env_key
    return DEFAULT_YOUTUBE_API_KEY


# ---------------------------------------------------------------------------
# YouTube scraping
# ---------------------------------------------------------------------------
def _scrap_youtube(hunt: CreatorHunt, niche: str, min_subs: int, max_subs: int,
                    num_results: int, api_key: str, pays: str = "FR",
                    progress_cb: Callable[[CreatorHunt], None] | None = None) -> list[Creator]:
    """Cherche des chaînes YouTube dans une niche, dans une fourchette d'abonnés,
    et tente d'extraire leur mail public.

    `pays` : code ISO (FR, BE, CH, CA, LU, MC, MA, DZ, TN, SN, CI) ou "ALL"
    pour tous les francophones. Limite l'API YouTube via `regionCode` +
    `relevanceLanguage='fr'`, puis filtre les chaînes non-francophones via
    langdetect (si installé) sur titre + description.
    """
    pays = _normalize_pays(pays)
    try:
        from googleapiclient.discovery import build
    except ImportError as exc:
        raise RuntimeError(
            "La bibliothèque google-api-python-client n'est pas installée. "
            "Lance `pip install google-api-python-client` sur le serveur."
        ) from exc

    def log(msg: str) -> None:
        hunt.log.append(f"[{datetime.now().strftime('%H:%M:%S')}] {msg}")
        if len(hunt.log) > 200:
            hunt.log = hunt.log[-200:]
        hunt.save()
        if progress_cb:
            try: progress_cb(hunt)
            except Exception: pass

    youtube = build("youtube", "v3", developerKey=api_key)

    zone_label = "tous francophones" if pays == "ALL" else pays
    log(f"🔎 Recherche YouTube : '{niche}' ({min_subs}-{max_subs} abonnés) — {zone_label}…")
    hunt.status = "searching"
    hunt.save()

    channels: list[dict] = []
    next_page_token: str | None = None
    page = 1
    max_pages = 10

    while len(channels) < num_results and page <= max_pages:
        params = {
            "q": niche,
            "part": "id,snippet",
            "type": "channel",
            "maxResults": 50,
            "relevanceLanguage": "fr",
        }
        if pays != "ALL":
            params["regionCode"] = pays
        if next_page_token:
            params["pageToken"] = next_page_token

        try:
            response = youtube.search().list(**params).execute()
        except Exception as exc:
            log(f"⚠️ Erreur API page {page} : {exc}")
            break

        raw_count = len(response.get("items", []))
        page_new = 0

        for item in response.get("items", []):
            if "channelId" not in item.get("id", {}):
                continue
            channel_id = item["id"]["channelId"]
            try:
                ch_resp = youtube.channels().list(
                    id=channel_id,
                    part="statistics,snippet,brandingSettings,contentDetails",
                ).execute()
                if not ch_resp.get("items"):
                    continue
                info = ch_resp["items"][0]
                sub_count = int(info["statistics"].get("subscriberCount", 0))
                if min_subs <= sub_count <= max_subs:
                    title = info["snippet"]["title"]
                    desc = info["snippet"].get("description", "")
                    # Filtre langue : si le titre + description ne sont pas
                    # en français, on écarte (langdetect ; en cas d'erreur
                    # ou de lib absente, on garde).
                    if not _is_french_text(f"{title} {desc}"):
                        continue
                    channels.append({
                        "id": channel_id,
                        "title": title,
                        "subs": sub_count,
                        "description": desc,
                        "country": info["snippet"].get("country", ""),
                        "created": info["snippet"].get("publishedAt", ""),
                        "video_count": int(info["statistics"].get("videoCount", 0)),
                    })
                    page_new += 1
                if len(channels) >= num_results:
                    break
            except Exception as exc:
                log(f"  ⚠️ Récup chaîne KO : {exc}")

        log(f"📄 Page {page}/{max_pages} : {raw_count} brute(s), +{page_new} retenu(s) (total {len(channels)})")

        next_page_token = response.get("nextPageToken")
        if not next_page_token:
            log("📄 Plus de pages disponibles dans cette niche.")
            break
        page += 1
        time.sleep(random.uniform(0.4, 1.2))

    hunt.stats["candidats"] = len(channels)
    log(f"✅ {len(channels)} chaîne(s) dans la fourchette.")

    # ----- Phase d'enrichissement : extraction des mails -----
    hunt.status = "enriching"
    hunt.save()
    results: list[Creator] = []
    total = max(len(channels), 1)

    for i, ch in enumerate(channels):
        log(f"--- {i+1}/{len(channels)} : {ch['title']} ({ch['subs']} abonnés)")
        emails: set[str] = set()
        externals: set[str] = set()

        # 1) Description de la chaîne
        for em in re.findall(EMAIL_REGEX, ch["description"]):
            emails.add(em.lower())

        # 2) Page /about de la chaîne
        try:
            about_url = f"https://www.youtube.com/channel/{ch['id']}/about"
            r = requests.get(about_url, headers=UA_WEB, timeout=10)
            html = r.text
            for em in re.findall(EMAIL_REGEX, html):
                emails.add(em.lower())
            for url in re.findall(r'https?://[^\s"\'<>\\]+', html):
                low = url.lower()
                for dom in BIO_DOMAINS:
                    if dom in low:
                        externals.add(url.split("\\")[0])
                        break
            log(f"  ✓ /about scrapé — {len(externals)} lien(s) bio")
        except Exception as exc:
            log(f"  ⚠️ /about KO : {exc}")

        # 3) Suivre chaque lien bio (linktree, etc.)
        for link in list(externals)[:5]:
            try:
                time.sleep(random.uniform(1.5, 3.5))
                r = requests.get(link, headers=UA_WEB, timeout=10)
                new = re.findall(EMAIL_REGEX, r.text)
                for em in new:
                    emails.add(em.lower())
                log(f"  ✓ {link[:60]} → {len(new)} mail(s)")
            except Exception as exc:
                log(f"  ⚠️ {link[:60]} KO : {exc}")

        # 4) Description des 3 dernières vidéos
        try:
            vid_resp = youtube.search().list(
                part="id", channelId=ch["id"],
                maxResults=3, order="date", type="video",
            ).execute()
            for v in vid_resp.get("items", []):
                vid = v["id"].get("videoId")
                if not vid:
                    continue
                vid_info = youtube.videos().list(part="snippet", id=vid).execute()
                if vid_info.get("items"):
                    desc = vid_info["items"][0]["snippet"].get("description", "")
                    for em in re.findall(EMAIL_REGEX, desc):
                        emails.add(em.lower())
        except Exception as exc:
            log(f"  ⚠️ Vidéos KO : {exc}")

        emails_clean = _clean_emails(emails)
        if emails_clean:
            log(f"  📧 {len(emails_clean)} mail(s) trouvé(s) : {', '.join(emails_clean[:3])}")
        else:
            log("  📧 aucun mail trouvé")

        creator = Creator(
            platform="youtube",
            handle=ch["id"],
            name=ch["title"],
            subscribers=ch["subs"],
            url=f"https://www.youtube.com/channel/{ch['id']}",
            email=emails_clean[0] if emails_clean else "",
            emails_extra=emails_clean[1:] if len(emails_clean) > 1 else [],
            description=ch["description"][:500],
            country=ch.get("country", ""),
            video_count=ch.get("video_count", 0),
            created_at=ch.get("created", ""),
            external_links=sorted(externals),
        )
        results.append(creator)

        hunt.creators = [c.to_dict() for c in results]
        hunt.progress = int((i + 1) / total * 100)
        hunt.stats["traites"] = i + 1
        hunt.stats["retenus"] = len(results)
        hunt.stats["avec_mail"] = sum(1 for c in results if c.email)
        hunt.save()
        if progress_cb:
            try: progress_cb(hunt)
            except Exception: pass

        time.sleep(random.uniform(1.5, 4.0))

    return results


# ---------------------------------------------------------------------------
# Instagram scraping (fragile — Meta bloque souvent)
# ---------------------------------------------------------------------------
def _scrap_instagram(hunt: CreatorHunt, niche: str, min_subs: int, max_subs: int,
                      num_results: int, login: str, password: str,
                      progress_cb: Callable[[CreatorHunt], None] | None) -> list[Creator]:
    try:
        from instagrapi import Client
    except ImportError as exc:
        raise RuntimeError(
            "La bibliothèque instagrapi n'est pas installée. "
            "Lance `pip install instagrapi` sur le serveur."
        ) from exc

    def log(msg: str) -> None:
        hunt.log.append(f"[{datetime.now().strftime('%H:%M:%S')}] {msg}")
        if len(hunt.log) > 200:
            hunt.log = hunt.log[-200:]
        hunt.save()
        if progress_cb:
            try: progress_cb(hunt)
            except Exception: pass

    if not login or not password:
        raise RuntimeError("Login et mot de passe Instagram requis.")

    hunt.status = "searching"
    hunt.save()
    log(f"🔎 Connexion à Instagram en tant que {login}…")
    cl = Client()
    cl.login(login, password)

    log(f"🔎 Recherche hashtag #{niche}…")
    tags = cl.hashtag_search(niche)
    accounts: list[tuple] = []
    seen: set[str] = set()
    for tag in tags[:5]:
        try:
            tag_id = tag.id
            feed = cl.hashtag_medias_recent(tag.name, amount=40)
            for media in feed:
                uid = str(media.user.pk)
                if uid in seen:
                    continue
                seen.add(uid)
                try:
                    info = cl.user_info(media.user.pk).dict()
                except Exception:
                    continue
                followers = int(info.get("follower_count", 0))
                if min_subs <= followers <= max_subs:
                    accounts.append((uid, info.get("username", ""), followers, info))
                if len(accounts) >= num_results:
                    break
            if len(accounts) >= num_results:
                break
        except Exception as exc:
            log(f"  ⚠️ tag KO : {exc}")
        time.sleep(random.uniform(2, 5))

    hunt.stats["candidats"] = len(accounts)
    log(f"✅ {len(accounts)} compte(s) dans la fourchette.")

    hunt.status = "enriching"
    hunt.save()
    results: list[Creator] = []
    total = max(len(accounts), 1)
    for i, (uid, username, followers, info) in enumerate(accounts):
        bio = info.get("biography", "") or ""
        emails: set[str] = set()
        for em in re.findall(EMAIL_REGEX, bio):
            emails.add(em.lower())
        for em in re.findall(EMAIL_REGEX, info.get("public_email", "") or ""):
            emails.add(em.lower())
        emails_clean = _clean_emails(emails)
        creator = Creator(
            platform="instagram",
            handle=username,
            name=info.get("full_name", username),
            subscribers=followers,
            url=f"https://www.instagram.com/{username}/",
            email=emails_clean[0] if emails_clean else "",
            emails_extra=emails_clean[1:] if len(emails_clean) > 1 else [],
            description=bio[:500],
        )
        results.append(creator)
        hunt.creators = [c.to_dict() for c in results]
        hunt.progress = int((i + 1) / total * 100)
        hunt.stats["traites"] = i + 1
        hunt.stats["retenus"] = len(results)
        hunt.stats["avec_mail"] = sum(1 for c in results if c.email)
        hunt.save()
        if progress_cb:
            try: progress_cb(hunt)
            except Exception: pass
        time.sleep(random.uniform(1.5, 3.5))

    return results


# ---------------------------------------------------------------------------
# Facebook scraping (très fragile)
# ---------------------------------------------------------------------------
def _scrap_facebook(hunt: CreatorHunt, niche: str, min_subs: int, max_subs: int,
                     num_results: int,
                     progress_cb: Callable[[CreatorHunt], None] | None) -> list[Creator]:
    try:
        from facebook_scraper import get_profile, get_posts  # type: ignore
    except ImportError as exc:
        raise RuntimeError(
            "La bibliothèque facebook-scraper n'est pas installée. "
            "Lance `pip install facebook-scraper` sur le serveur."
        ) from exc

    def log(msg: str) -> None:
        hunt.log.append(f"[{datetime.now().strftime('%H:%M:%S')}] {msg}")
        if len(hunt.log) > 200:
            hunt.log = hunt.log[-200:]
        hunt.save()
        if progress_cb:
            try: progress_cb(hunt)
            except Exception: pass

    hunt.status = "searching"
    hunt.save()
    log(f"🔎 Recherche Facebook : {niche}…")

    pages: list[dict] = []
    try:
        for post in get_posts(niche, pages=10):
            uname = post.get("username") or post.get("user_id")
            if not uname:
                continue
            try:
                info = get_profile(uname)
                likes = int(info.get("likes") or info.get("Friends") or 0)
                if min_subs <= likes <= max_subs:
                    pages.append({"id": uname, "name": info.get("Name", uname),
                                  "likes": likes, "about": info.get("about", "")})
                if len(pages) >= num_results:
                    break
            except Exception:
                continue
            time.sleep(random.uniform(2, 4))
    except Exception as exc:
        log(f"  ⚠️ Recherche Facebook KO : {exc}")

    hunt.stats["candidats"] = len(pages)
    log(f"✅ {len(pages)} page(s).")

    hunt.status = "enriching"
    hunt.save()
    results: list[Creator] = []
    total = max(len(pages), 1)
    for i, page in enumerate(pages):
        emails: set[str] = set()
        for em in re.findall(EMAIL_REGEX, page.get("about", "") or ""):
            emails.add(em.lower())
        emails_clean = _clean_emails(emails)
        creator = Creator(
            platform="facebook",
            handle=page["id"],
            name=page["name"],
            subscribers=page["likes"],
            url=f"https://www.facebook.com/{page['id']}",
            email=emails_clean[0] if emails_clean else "",
            emails_extra=emails_clean[1:] if len(emails_clean) > 1 else [],
            description=(page.get("about") or "")[:500],
        )
        results.append(creator)
        hunt.creators = [c.to_dict() for c in results]
        hunt.progress = int((i + 1) / total * 100)
        hunt.stats["traites"] = i + 1
        hunt.stats["retenus"] = len(results)
        hunt.stats["avec_mail"] = sum(1 for c in results if c.email)
        hunt.save()
        if progress_cb:
            try: progress_cb(hunt)
            except Exception: pass

    return results


# ---------------------------------------------------------------------------
# Orchestration : threads, list, export
# ---------------------------------------------------------------------------
_RUNNING: dict[str, threading.Thread] = {}


def is_running(hunt_id: str) -> bool:
    t = _RUNNING.get(hunt_id)
    return bool(t and t.is_alive())


def list_hunts(limit: int = 20) -> list[dict]:
    """Liste les chasses connues localement, plus récentes en premier.

    Requalifie au passage les chasses zombies (interrompues par un
    redémarrage du serveur).
    """
    ensure_dirs()
    from .hunt_zombies import reconcile_hunt_file
    items: list[dict] = []
    for p in HUNTS_DIR.glob("*.json"):
        try:
            running = is_running(p.stem)
            data = reconcile_hunt_file(p, is_running=running)
            if not data:
                continue
            items.append({
                "id": data.get("id"),
                "label": data.get("label"),
                "created_at": data.get("created_at"),
                "status": data.get("status"),
                "progress": data.get("progress"),
                "stats": data.get("stats") or {},
                "filters": data.get("filters") or {},
                "running": running,
                "creators_count": len(data.get("creators") or []),
            })
        except Exception:
            continue
    items.sort(key=lambda x: x.get("created_at") or "", reverse=True)
    return items[:limit]


def _build_label(platform: str, niche: str, min_subs: int, max_subs: int) -> str:
    p = {"youtube": "YouTube", "instagram": "Instagram",
         "facebook": "Facebook"}.get(platform, platform)
    n = (niche or "").strip() or "tous"
    return f"{p} · {n} · {min_subs:,}–{max_subs:,} abonnés".replace(",", " ")


def start_hunt(platform: str, niche: str, min_subs: int, max_subs: int,
               num_results: int = 50,
               youtube_api_key: str | None = None,
               instagram_login: str | None = None,
               instagram_password: str | None = None,
               pays: str = "FR",
               progress_cb: Callable[[CreatorHunt], None] | None = None,
               ) -> CreatorHunt:
    """Lance une chasse aux créateurs en arrière-plan.

    `pays` : code ISO francophone (FR par défaut). Voir FRANCOPHONE_COUNTRIES.
    "ALL" = tous francophones (pas de regionCode, filtre langue uniquement).
    """
    ensure_dirs()
    if platform not in ("youtube", "instagram", "facebook"):
        raise ValueError(f"Plateforme inconnue : {platform}")
    if not niche or not niche.strip():
        raise ValueError("La niche / mot-clé est requise.")
    if min_subs < 0:
        min_subs = 0
    if max_subs <= min_subs:
        max_subs = min_subs + 1
    if num_results <= 0:
        num_results = 50
    if num_results > 500:
        num_results = 500
    pays = _normalize_pays(pays)

    hunt_id = uuid.uuid4().hex[:12]
    label = _build_label(platform, niche, min_subs, max_subs)
    hunt = CreatorHunt(
        id=hunt_id,
        label=label,
        created_at=datetime.now(timezone.utc).isoformat(timespec="seconds"),
        status="pending",
        progress=0,
        filters={
            "platform": platform, "niche": niche,
            "min_subs": min_subs, "max_subs": max_subs,
            "num_results": num_results, "pays": pays,
        },
    )
    hunt.save()

    def _thread():
        try:
            if platform == "youtube":
                key = _get_youtube_api_key(youtube_api_key)
                _scrap_youtube(hunt, niche, min_subs, max_subs, num_results,
                               key, pays=pays, progress_cb=progress_cb)
            elif platform == "instagram":
                _scrap_instagram(hunt, niche, min_subs, max_subs, num_results,
                                 instagram_login or "", instagram_password or "",
                                 progress_cb)
            elif platform == "facebook":
                _scrap_facebook(hunt, niche, min_subs, max_subs, num_results,
                                progress_cb)
            hunt.status = "done"
            hunt.progress = 100
            hunt.save()
        except Exception as exc:
            logger.exception("creator hunt %s crashed", hunt.id)
            hunt.status = "error"
            hunt.error = str(exc)
            hunt.save()

    t = threading.Thread(target=_thread, daemon=True,
                         name=f"chasseur-c-{hunt_id}")
    _RUNNING[hunt_id] = t
    t.start()
    return hunt


def export_csv(hunt_id: str) -> dict:
    """Exporte les créateurs d'une chasse en CSV."""
    h = CreatorHunt.load(hunt_id)
    if not h:
        return {"ok": False, "error": "chasse introuvable"}
    ensure_dirs()
    out = EXPORTS_DIR / f"createurs_{h.id}.csv"
    rows = h.creators or []
    with out.open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow([
            "Plateforme", "Nom", "Abonnés", "Email principal",
            "Emails secondaires", "URL", "Pays", "Vidéos",
            "Date création", "Liens externes",
        ])
        for c in rows:
            w.writerow([
                c.get("platform", ""),
                c.get("name", ""),
                c.get("subscribers", 0),
                c.get("email", ""),
                "; ".join(c.get("emails_extra", []) or []),
                c.get("url", ""),
                c.get("country", ""),
                c.get("video_count", 0),
                c.get("created_at", ""),
                "; ".join(c.get("external_links", []) or []),
            ])
    return {"ok": True, "path": str(out), "rows": len(rows)}


def export_xlsx(hunt_id: str) -> dict:
    """Exporte les créateurs d'une chasse en fichier Excel (.xlsx)."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError as exc:
        return {"ok": False,
                "error": "openpyxl manquant — `pip install openpyxl`"}
    h = CreatorHunt.load(hunt_id)
    if not h:
        return {"ok": False, "error": "chasse introuvable"}
    ensure_dirs()
    out = EXPORTS_DIR / f"createurs_{h.id}.xlsx"
    rows = h.creators or []

    wb = Workbook()
    ws = wb.active
    ws.title = "Créateurs"
    headers = ["Plateforme", "Nom", "Abonnés", "Email principal",
               "Emails secondaires", "URL", "Pays", "Vidéos",
               "Date création", "Liens externes"]
    ws.append(headers)

    # Mise en forme de l'entête (gras blanc sur fond orange Triskell)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="EA580C")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left", vertical="center")

    for c in rows:
        ws.append([
            c.get("platform", ""),
            c.get("name", ""),
            c.get("subscribers", 0),
            c.get("email", ""),
            "; ".join(c.get("emails_extra", []) or []),
            c.get("url", ""),
            c.get("country", ""),
            c.get("video_count", 0),
            c.get("created_at", ""),
            "; ".join(c.get("external_links", []) or []),
        ])

    # Largeurs de colonnes ajustées au contenu attendu
    widths = [12, 32, 12, 30, 30, 40, 10, 10, 18, 40]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w

    # Fige la première ligne pour qu'elle reste visible au scroll
    ws.freeze_panes = "A2"

    wb.save(out)
    return {"ok": True, "path": str(out), "rows": len(rows)}


def delete_hunt(hunt_id: str) -> dict:
    h = CreatorHunt.load(hunt_id)
    if not h:
        return {"ok": False, "error": "chasse introuvable"}
    try:
        h.path.unlink(missing_ok=True)
    except Exception as exc:
        return {"ok": False, "error": str(exc)}
    return {"ok": True}


def push_to_prospects(hunt_id: str) -> dict:
    """Pousse les créateurs d'une chasse vers la base partagée (table
    `prospects`) — la même base qu'Obélisk, l'Auto-Pilote et la vue
    "Tous les prospects". Upsert dédoublonné par email / URL.

    Seuls les créateurs AVEC email sont poussés : sans mail, ni
    l'Auto-Pilote ni les campagnes ne peuvent rien en faire.

    Renvoie {ok, backend, pushed, created, merged, total}.
    """
    h = CreatorHunt.load(hunt_id)
    if not h:
        return {"ok": False, "error": "chasse introuvable"}
    if not h.creators:
        return {"ok": False, "error": "aucun créateur à pousser"}

    try:
        from triskell_core.prospect.core.crm import get_crm
        from triskell_core.prospect.core.prospect import (
            Prospect as CoreProspect, Source,
        )
    except ImportError as exc:
        return {"ok": False, "error":
                f"triskell_core absent — impossible de pousser ({exc})"}

    try:
        crm = get_crm()
    except Exception as exc:
        return {"ok": False, "error": f"connexion CRM impossible : {exc}"}
    backend = "remote" if crm.__class__.__name__ == "RemoteCRM" else "local"

    niche = (h.filters or {}).get("niche") or ""

    context_by_platform = {
        "youtube": "profil YouTube (description / page À propos)",
        "instagram": "bio Instagram",
        "facebook": "page Facebook",
    }

    core_prospects: list[CoreProspect] = []
    for c in h.creators:
        email = (c.get("email") or "").strip()
        if not email:
            continue
        platform = (c.get("platform") or "").strip() or "youtube"
        all_emails = [email] + [e for e in (c.get("emails_extra") or []) if e]
        emails_meta = [{
            "email": e,
            "source": platform,
            "source_id": (c.get("handle") or "").strip(),
            "url": (c.get("url") or "").strip(),
            "context": context_by_platform.get(platform, "profil créateur"),
            "found_at": "",
        } for e in all_emails]
        try:
            subs = int(c.get("subscribers") or 0)
        except (TypeError, ValueError):
            subs = 0
        cp = CoreProspect(
            name=(c.get("name") or "").strip(),
            handle=(c.get("handle") or "").strip(),
            emails=all_emails,
            emails_meta=emails_meta,
            country=(c.get("country") or "").strip(),
            industry=niche,
            description=(c.get("description") or "").strip()[:500],
            language="fr",
            subscribers=subs or None,
            platform_url=(c.get("url") or "").strip(),
            other_urls=[u for u in (c.get("external_links") or []) if u][:5],
            tags=["chasseur_createurs", platform],
            sources=[Source(
                name=platform,
                source_id=(c.get("handle") or "").strip(),
                url=(c.get("url") or "").strip(),
            )],
            status="new",
        )
        core_prospects.append(cp)

    if not core_prospects:
        return {"ok": False, "error":
                "aucun créateur avec mail à pousser (mail requis)"}

    try:
        result = crm.upsert_many(core_prospects)
        if hasattr(crm, "save"):
            try: crm.save()
            except Exception: pass
        return {
            "ok": True,
            "backend":  backend,
            "created":  int(result.get("created") or 0),
            "merged":   int(result.get("merged") or 0),
            "total":    int(result.get("total") or 0),
            "pushed":   len(core_prospects),
        }
    except Exception as exc:
        return {"ok": False, "error": f"upsert échoué : {exc}"}
