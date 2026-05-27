"""
Export Excel : une seule colonne "Email", une ligne par adresse.
"""

from datetime import datetime
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Font


def build_filename() -> str:
    """Nom de fichier au format demandé : emails_AAAA-MM-JJ_HHmm.xlsx"""
    return f"emails_{datetime.now().strftime('%Y-%m-%d_%H%M')}.xlsx"


def export_to_excel(emails: Iterable[str], output_path: Path) -> Path:
    """Écrit le fichier Excel et retourne le chemin du fichier généré."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Emails"

    # En-tête.
    ws["A1"] = "Email"
    ws["A1"].font = Font(bold=True)
    ws.column_dimensions["A"].width = 50

    # Lignes : une par email, trié alphabétiquement pour la lisibilité.
    for row, email in enumerate(sorted(set(emails)), start=2):
        ws.cell(row=row, column=1, value=email)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path


def read_reference_emails(file_path: Path) -> set[str]:
    """
    Lit un fichier Excel précédent pour récupérer les emails à exclure du run.
    On lit la colonne A en sautant la première ligne (en-tête).
    """
    from openpyxl import load_workbook

    if not file_path.exists():
        return set()

    wb = load_workbook(file_path, read_only=True, data_only=True)
    ws = wb.active
    found: set[str] = set()
    for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
        if row_idx == 0:
            continue
        if row and row[0]:
            found.add(str(row[0]).strip().lower())
    wb.close()
    return found
