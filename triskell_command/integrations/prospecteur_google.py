"""Prospecteur Google Places — découvre des entreprises locales via l'API
Google Places (Text Search v1) puis tente d'extraire leurs mails publics en
scrapant leur site web.

Adapté depuis le programme desktop `prospecteur-google/main.py` (Tkinter). On
réutilise l'architecture "hunt en arrière-plan" déjà en place pour le
Chasseur PME et le Chasseur Créateur : chaque recherche tourne dans un
thread daemon, persiste son état en JSON, et l'UI poll toutes les 2 secondes.

Pourquoi c'est utile :
  - Beaucoup d'entreprises locales sont sur Google Maps mais n'ont pas de
    site web → cibles idéales Triskell (vendre un site).
  - Pour celles qui en ont un, on récupère leur mail public en scrapant les
    pages habituelles (home, contact, mentions légales, à propos).

L'utilisateur peut filtrer pour ne garder que les entreprises SANS site web
(prospects "site neuf à vendre").
"""
from __future__ import annotations

import csv
import json
import logging
import os
import random
import re
import threading
import time
import uuid
from dataclasses import asdict, dataclass, field
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Callable

import requests

logger = logging.getLogger(__name__)


PROSP_DIR = Path.home() / ".triskell-command" / "prospecteur_google"
HUNTS_DIR = PROSP_DIR / "hunts"
EXPORTS_DIR = PROSP_DIR / "exports"

# Clé API Google Places par défaut (issue du programme desktop d'origine).
# Surchargeable via la variable d'env GOOGLE_PLACES_API_KEY ou dans le payload.
DEFAULT_GOOGLE_PLACES_KEY = "AIzaSyAhOg-mJY2TTOnl4mUKi_FCRt2b__jHd2I"

PLACES_URL = "https://places.googleapis.com/v1/places:searchText"
PLACES_FIELDS = (
    "places.displayName,places.formattedAddress,"
    "places.internationalPhoneNumber,places.websiteUri,"
    "places.businessStatus,nextPageToken"
)

EMAIL_REGEX = r"[\w\.\-+]+@[\w\.\-]+\.\w+"

UA_WEB = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/120.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "fr-FR,fr;q=0.9,en;q=0.8",
}

# Mails techniques / faux positifs à jeter
EMAIL_BLACKLIST = [
    "@example.", "noreply", "no-reply", "@wix.com", "@sentry",
    "@wordpress", "godaddy", "domain.com", "@sentry.io",
    "@cloudfront", ".png", ".jpg", "@google.", "@gstatic",
    "@bootstrapcdn", "@jquery", "@w3.org",
]


def ensure_dirs() -> None:
    HUNTS_DIR.mkdir(parents=True, exist_ok=True)
    EXPORTS_DIR.mkdir(parents=True, exist_ok=True)


def _get_api_key(payload_key: str | None = None) -> str:
    if payload_key and payload_key.strip():
        return payload_key.strip()
    env_key = (os.environ.get("GOOGLE_PLACES_API_KEY") or "").strip()
    if env_key:
        return env_key
    return DEFAULT_GOOGLE_PLACES_KEY


# ---------------------------------------------------------------------------
# Modèles
# ---------------------------------------------------------------------------
@dataclass
class GoogleProspect:
    """Une entreprise trouvée via Google Places."""
    name: str
    address: str = ""
    phone: str = ""
    website: str = ""
    email: str = ""               # mail principal trouvé sur le site
    emails_extra: list[str] = field(default_factory=list)
    has_website: bool = False
    status: str = ""              # OPERATIONAL / CLOSED_TEMPORARILY / etc.

    def to_dict(self) -> dict:
        return asdict(self)


@dataclass
class ProspectHunt:
    id: str
    label: str
    created_at: str
    status: str = "pending"   # pending / searching / enriching / done / error
    progress: int = 0          # 0-100
    log: list[str] = field(default_factory=list)
    filters: dict = field(default_factory=dict)
    stats: dict = field(default_factory=dict)
    prospects: list[dict] = field(default_factory=list)
    error: str = ""

    @property
    def path(self) -> Path:
        return HUNTS_DIR / f"{self.id}.json"

    def save(self) -> None:
        ensure_dirs()
        self.path.write_text(
            json.dumps(asdict(self), ensure_ascii=False, indent=2),
            encoding="utf-8",
        )

    @classmethod
    def load(cls, hunt_id: str) -> "ProspectHunt | None":
        p = HUNTS_DIR / f"{hunt_id}.json"
        if not p.exists():
            return None
        try:
            data = json.loads(p.read_text(encoding="utf-8"))
            return cls(**data)
        except Exception as exc:
            logger.warning("prospect hunt.load failed %s: %s", hunt_id, exc)
            return None


# ---------------------------------------------------------------------------
# Helpers scraping site
# ---------------------------------------------------------------------------
def _scrape_site_for_emails(url: str) -> set[str]:
    """Scrape les pages habituelles d'un site pour en extraire des mails."""
    emails: set[str] = set()
    if not url:
        return emails
    if not url.startswith("http"):
        url = "https://" + url
    base = url.rstrip("/")
    pages = [
        url,
        f"{base}/contact",
        f"{base}/contact.html",
        f"{base}/contact.php",
        f"{base}/a-propos",
        f"{base}/about",
        f"{base}/mentions-legales",
    ]
    for page in pages:
        try:
            r = requests.get(page, headers=UA_WEB, timeout=8)
            if r.status_code == 200:
                for em in re.findall(EMAIL_REGEX, r.text):
                    em_low = em.lower()
                    if not any(b in em_low for b in EMAIL_BLACKLIST):
                        emails.add(em_low)
        except Exception:
            pass
        time.sleep(random.uniform(0.3, 0.6))
    return emails


# ---------------------------------------------------------------------------
# Orchestration : run, threads, list, export
# ---------------------------------------------------------------------------
_RUNNING: dict[str, threading.Thread] = {}


def is_running(hunt_id: str) -> bool:
    t = _RUNNING.get(hunt_id)
    return bool(t and t.is_alive())


def list_hunts(limit: int = 20) -> list[dict]:
    ensure_dirs()
    items: list[dict] = []
    for p in HUNTS_DIR.glob("*.json"):
        try:
            data = json.loads(p.read_text(encoding="utf-8"))
            items.append({
                "id": data.get("id"),
                "label": data.get("label"),
                "created_at": data.get("created_at"),
                "status": data.get("status"),
                "progress": data.get("progress"),
                "stats": data.get("stats") or {},
                "filters": data.get("filters") or {},
                "running": is_running(data.get("id") or ""),
            })
        except Exception:
            continue
    items.sort(key=lambda x: x.get("created_at") or "", reverse=True)
    return items[:limit]


def _build_label(metier: str, zone: str, only_no_site: bool) -> str:
    m = (metier or "").strip() or "tous métiers"
    z = (zone or "").strip() or "France"
    suffix = "  ·  sans site" if only_no_site else ""
    return f"{m} — {z}{suffix}"


def start_hunt(metier: str, zone: str, num_results: int = 60,
               only_no_site: bool = False,
               api_key: str | None = None,
               progress_cb: Callable[[ProspectHunt], None] | None = None,
               ) -> ProspectHunt:
    """Lance une chasse Google Places en arrière-plan."""
    ensure_dirs()
    if not metier or not metier.strip():
        raise ValueError("Indique un métier ou un type d'entreprise.")
    if not zone or not zone.strip():
        raise ValueError("Indique une zone géographique.")
    if num_results <= 0:
        num_results = 60
    if num_results > 200:
        num_results = 200

    hunt_id = uuid.uuid4().hex[:12]
    hunt = ProspectHunt(
        id=hunt_id,
        label=_build_label(metier, zone, only_no_site),
        created_at=datetime.now(timezone.utc).isoformat(timespec="seconds"),
        status="pending",
        progress=0,
        filters={
            "metier": metier, "zone": zone,
            "num_results": num_results, "only_no_site": only_no_site,
        },
    )
    hunt.save()

    def _thread():
        try:
            _run_hunt(hunt, metier=metier, zone=zone,
                      num_results=num_results, only_no_site=only_no_site,
                      api_key=_get_api_key(api_key),
                      progress_cb=progress_cb)
            hunt.status = "done"
            hunt.progress = 100
            hunt.save()
        except Exception as exc:
            logger.exception("prospect hunt %s crashed", hunt.id)
            hunt.status = "error"
            hunt.error = str(exc)
            hunt.save()

    t = threading.Thread(target=_thread, daemon=True,
                         name=f"prospecteur-{hunt_id}")
    _RUNNING[hunt_id] = t
    t.start()
    return hunt


def _run_hunt(hunt: ProspectHunt, metier: str, zone: str, num_results: int,
              only_no_site: bool, api_key: str,
              progress_cb: Callable[[ProspectHunt], None] | None) -> None:
    def log(msg: str) -> None:
        hunt.log.append(f"[{datetime.now().strftime('%H:%M:%S')}] {msg}")
        if len(hunt.log) > 200:
            hunt.log = hunt.log[-200:]
        hunt.save()
        if progress_cb:
            try: progress_cb(hunt)
            except Exception: pass

    hunt.status = "searching"
    hunt.save()
    log(f"🔎 Recherche Google Places : '{metier}' à '{zone}' (max {num_results})")

    headers = {
        "Content-Type": "application/json",
        "X-Goog-Api-Key": api_key,
        "X-Goog-FieldMask": PLACES_FIELDS,
    }
    body: dict[str, Any] = {"textQuery": f"{metier} {zone}", "pageSize": 20}

    all_places: list[dict] = []
    page_token: str | None = None
    while len(all_places) < num_results:
        if page_token:
            body["pageToken"] = page_token
        try:
            r = requests.post(PLACES_URL, headers=headers, json=body, timeout=15)
            if r.status_code != 200:
                log(f"⚠️ Erreur API ({r.status_code}) : {r.text[:300]}")
                break
            data = r.json()
            places = data.get("places", []) or []
            all_places.extend(places)
            log(f"📄 +{len(places)} entreprise(s) (total {len(all_places)})")
            page_token = data.get("nextPageToken")
            if not page_token:
                break
            time.sleep(2)
        except Exception as exc:
            log(f"⚠️ Erreur réseau : {exc}")
            break

    all_places = all_places[:num_results]
    hunt.stats["candidats"] = len(all_places)
    log(f"✅ {len(all_places)} entreprise(s) à analyser.")

    # ----- Phase enrichissement : scraping mails -----
    hunt.status = "enriching"
    hunt.save()
    results: list[GoogleProspect] = []
    total = max(len(all_places), 1)

    for i, place in enumerate(all_places):
        name = (place.get("displayName") or {}).get("text") or "Inconnu"
        address = place.get("formattedAddress", "") or ""
        phone = place.get("internationalPhoneNumber", "") or ""
        website = place.get("websiteUri", "") or ""
        status = place.get("businessStatus", "") or ""

        log(f"--- {i+1}/{len(all_places)} : {name}")
        log(f"  📍 {address}")
        if phone:
            log(f"  📞 {phone}")

        emails: list[str] = []
        if website:
            log(f"  🌐 {website}")
            found = _scrape_site_for_emails(website)
            emails = sorted(found)
            if emails:
                log(f"  📧 {len(emails)} mail(s) : {', '.join(emails[:3])}")
            else:
                log("  📧 aucun mail trouvé sur le site")
        else:
            log("  ⚠️ Pas de site web → cible Triskell idéale")

        prospect = GoogleProspect(
            name=name, address=address, phone=phone, website=website,
            email=emails[0] if emails else "",
            emails_extra=emails[1:] if len(emails) > 1 else [],
            has_website=bool(website),
            status=status,
        )
        results.append(prospect)

        # Si only_no_site, on ne garde que celles sans site dans la sortie,
        # mais on continue à logger toutes les boîtes pour la transparence.
        retained = ([p for p in results if not p.has_website]
                    if only_no_site else results)

        hunt.prospects = [p.to_dict() for p in retained]
        hunt.progress = int((i + 1) / total * 100)
        hunt.stats["traites"] = i + 1
        hunt.stats["retenus"] = len(retained)
        hunt.stats["avec_mail"] = sum(1 for p in retained if p.email)
        hunt.stats["sans_site"] = sum(1 for p in results if not p.has_website)
        hunt.save()
        if progress_cb:
            try: progress_cb(hunt)
            except Exception: pass

        time.sleep(random.uniform(0.3, 0.8))

    no_site = sum(1 for p in results if not p.has_website)
    log(f"🎉 Terminé — {len(results)} entreprise(s) analysées · {no_site} sans site web")


def export_csv(hunt_id: str) -> dict:
    h = ProspectHunt.load(hunt_id)
    if not h:
        return {"ok": False, "error": "chasse introuvable"}
    ensure_dirs()
    out = EXPORTS_DIR / f"prospects_google_{h.id}.csv"
    rows = h.prospects or []
    with out.open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow([
            "Nom", "Adresse", "Téléphone", "Site web",
            "Email principal", "Emails secondaires",
            "A un site ?", "Statut",
        ])
        for p in rows:
            w.writerow([
                p.get("name", ""),
                p.get("address", ""),
                p.get("phone", ""),
                p.get("website", ""),
                p.get("email", ""),
                "; ".join(p.get("emails_extra", []) or []),
                "oui" if p.get("has_website") else "NON",
                p.get("status", ""),
            ])
    return {"ok": True, "path": str(out), "rows": len(rows)}


def export_xlsx(hunt_id: str) -> dict:
    """Exporte les entreprises d'une recherche en fichier Excel (.xlsx)."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return {"ok": False,
                "error": "openpyxl manquant — `pip install openpyxl`"}
    h = ProspectHunt.load(hunt_id)
    if not h:
        return {"ok": False, "error": "chasse introuvable"}
    ensure_dirs()
    out = EXPORTS_DIR / f"prospects_google_{h.id}.xlsx"
    rows = h.prospects or []

    wb = Workbook()
    ws = wb.active
    ws.title = "Entreprises"
    headers = ["Nom", "Adresse", "Téléphone", "Site web",
               "Email principal", "Emails secondaires",
               "A un site ?", "Statut"]
    ws.append(headers)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="EA580C")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left", vertical="center")

    # Couleur de fond pour mettre en évidence les boîtes SANS site
    no_site_fill = PatternFill("solid", fgColor="FEF3C7")  # ambre clair

    for p in rows:
        has_site = bool(p.get("has_website"))
        ws.append([
            p.get("name", ""),
            p.get("address", ""),
            p.get("phone", ""),
            p.get("website", ""),
            p.get("email", ""),
            "; ".join(p.get("emails_extra", []) or []),
            "oui" if has_site else "NON",
            p.get("status", ""),
        ])
        # Si pas de site, fond ambre clair pour toute la ligne
        if not has_site:
            for cell in ws[ws.max_row]:
                cell.fill = no_site_fill

    widths = [28, 38, 18, 32, 28, 28, 12, 18]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w

    ws.freeze_panes = "A2"

    wb.save(out)
    return {"ok": True, "path": str(out), "rows": len(rows)}


def delete_hunt(hunt_id: str) -> dict:
    h = ProspectHunt.load(hunt_id)
    if not h:
        return {"ok": False, "error": "chasse introuvable"}
    try:
        h.path.unlink(missing_ok=True)
    except Exception as exc:
        return {"ok": False, "error": str(exc)}
    return {"ok": True}
