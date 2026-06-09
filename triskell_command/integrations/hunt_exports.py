"""Fabricant d'exports Excel commun aux outils de chasse.

Avant : Chasseur Créateur et Prospecteur Google avaient chacun leur copie
quasi identique du même code openpyxl (entête orange Triskell, largeurs de
colonnes, ligne figée). Une seule implémentation désormais — un correctif
de style profite à tous les outils d'un coup.
"""
from __future__ import annotations

from pathlib import Path
from typing import Iterable, Sequence

# Couleurs maison
HEADER_BG = "EA580C"        # orange Triskell
HIGHLIGHT_BG = "FEF3C7"     # ambre clair (lignes à mettre en avant)


def write_xlsx(path: Path | str, *, sheet_title: str,
               headers: Sequence[str],
               rows: Iterable[Sequence],
               widths: Sequence[int],
               highlight: Iterable[bool] | None = None) -> int:
    """Écrit un fichier Excel stylé Triskell. Renvoie le nombre de lignes.

    - `rows`      : lignes de valeurs, dans l'ordre des `headers`.
    - `widths`    : largeur de chaque colonne.
    - `highlight` : optionnel, un booléen par ligne — True = fond ambre
                    (ex : entreprises SANS site chez le Prospecteur Google).

    Lève ImportError si openpyxl manque (l'appelant la traduit en message).
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title
    ws.append(list(headers))

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor=HEADER_BG)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left", vertical="center")

    highlight_fill = PatternFill("solid", fgColor=HIGHLIGHT_BG)
    flags = list(highlight) if highlight is not None else None

    count = 0
    for idx, row in enumerate(rows):
        ws.append(list(row))
        if flags is not None and idx < len(flags) and flags[idx]:
            for cell in ws[ws.max_row]:
                cell.fill = highlight_fill
        count += 1

    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w

    ws.freeze_panes = "A2"
    wb.save(str(path))
    return count
