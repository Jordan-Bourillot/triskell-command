"""Export Obelisk — génère un fichier Excel ou PDF à partir d'une liste
de prospects récupérée via repo.list_creators.

Deux formats supportés :
- xlsx : openpyxl, une feuille tableur avec toutes les colonnes utiles
- pdf  : reportlab, format paysage A4 avec ligne par prospect

Les deux renvoient des `bytes` que le serveur HTTP encode en base64 et
envoie au navigateur, qui déclenche le download.
"""
from __future__ import annotations

import io
import logging
from datetime import datetime
from typing import Iterable

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Helpers communs
# ---------------------------------------------------------------------------
COLUMNS = [
    # (clé interne, libellé colonne, largeur xlsx, largeur PDF (mm))
    ("name",         "Nom",            32, 38),
    ("handle",       "Pseudo",         20, 22),
    ("platform",     "Plateforme",     14, 18),
    ("emails",       "Email",          34, 50),
    ("phones",       "Téléphone",      18, 22),
    ("website",      "Site web",       38, 45),
    ("subscribers",  "Abonnés",        12, 14),
    ("country",      "Pays",            8, 10),
    ("city",         "Ville",          18, 20),
    ("status",       "Statut",         12, 14),
    ("score",        "Score",           8, 10),
    ("description",  "Description",    60, 70),
    ("platform_url", "URL profil",     40, 50),
    ("created_at",   "Trouvé le",      18, 22),
]


def _join_list(v) -> str:
    if not v:
        return ""
    if isinstance(v, list):
        return ", ".join(str(x) for x in v if x)
    return str(v)


def _platform_of(row: dict) -> str:
    """Déduit la plateforme depuis sources[0].name (ex: 'obelisk_youtube')
    ou depuis platform_url."""
    sources = row.get("sources") or []
    if sources and isinstance(sources, list):
        src = sources[0] if isinstance(sources[0], dict) else {}
        name = (src.get("name") or "").lower()
        for p in ("youtube", "twitch", "reddit", "instagram", "tiktok",
                  "linkedin", "bluesky", "mastodon", "dailymotion",
                  "kick", "github", "podcast"):
            if p in name:
                return p
    url = (row.get("platform_url") or "").lower()
    for p in ("youtube", "twitch", "reddit", "instagram", "tiktok",
              "linkedin", "bluesky", "mastodon", "dailymotion",
              "kick", "github"):
        if p in url:
            return p
    return ""


def _value_for(row: dict, key: str) -> str:
    if key == "platform":
        return _platform_of(row)
    if key in ("emails", "phones"):
        return _join_list(row.get(key))
    v = row.get(key)
    if v is None:
        return ""
    return str(v)


# ---------------------------------------------------------------------------
# Excel
# ---------------------------------------------------------------------------
def to_xlsx(rows: Iterable[dict], *, title: str = "Prospects Obelisk") -> bytes:
    """Renvoie le fichier .xlsx en bytes. Lisible Excel/Google Sheets/LibreOffice."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise RuntimeError(f"openpyxl manquant : {exc}")

    wb = Workbook()
    ws = wb.active
    ws.title = "Prospects"

    # En-tête
    header_fill = PatternFill("solid", fgColor="2D2D44")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    header_align = Alignment(horizontal="left", vertical="center", wrap_text=False)
    border = Border(
        bottom=Side(border_style="thin", color="C0C0C0"),
    )
    for col_idx, (_, label, width, _) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[1].height = 22

    # Données
    body_align = Alignment(vertical="top", wrap_text=True)
    for r_idx, row in enumerate(rows, start=2):
        for c_idx, (key, _, _, _) in enumerate(COLUMNS, start=1):
            v = _value_for(row, key)
            cell = ws.cell(row=r_idx, column=c_idx, value=v)
            cell.alignment = body_align
            cell.border = border

    # Fige la première ligne
    ws.freeze_panes = "A2"

    # Métadonnées
    wb.properties.title = title
    wb.properties.creator = "Triskell Command — Obelisk"
    wb.properties.created = datetime.now()

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# PDF
# ---------------------------------------------------------------------------
def to_pdf(rows: list[dict], *, title: str = "Prospects Obelisk") -> bytes:
    """Renvoie le PDF en bytes. Layout paysage A4, une ligne par prospect.

    Pour rester lisible, on prend un sous-ensemble de colonnes (les plus
    utiles pour un coup d'œil). Les colonnes longues (description, URL)
    sont tronquées.
    """
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import mm
        from reportlab.platypus import (
            SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
        )
        from reportlab.lib.enums import TA_LEFT
    except ImportError as exc:
        raise RuntimeError(
            f"reportlab manquant pour l'export PDF : {exc}. "
            f"Installe avec : pip install reportlab"
        )

    rows = list(rows)
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=landscape(A4),
        leftMargin=12 * mm,
        rightMargin=12 * mm,
        topMargin=14 * mm,
        bottomMargin=14 * mm,
        title=title,
        author="Triskell Command — Obelisk",
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "Title",
        parent=styles["Heading1"],
        fontSize=16,
        textColor=colors.HexColor("#2D2D44"),
        spaceAfter=4,
    )
    sub_style = ParagraphStyle(
        "Sub",
        parent=styles["Normal"],
        fontSize=9,
        textColor=colors.HexColor("#888"),
        spaceAfter=12,
    )
    cell_style = ParagraphStyle(
        "Cell",
        parent=styles["Normal"],
        fontSize=8,
        leading=10,
        alignment=TA_LEFT,
    )

    # Sous-ensemble de colonnes lisibles en paysage
    pdf_cols = [
        ("name",         "Nom",          34 * mm),
        ("handle",       "Pseudo",       22 * mm),
        ("platform",     "Plateforme",   18 * mm),
        ("emails",       "Email",        50 * mm),
        ("subscribers",  "Abonnés",      18 * mm),
        ("country",      "Pays",         12 * mm),
        ("website",      "Site web",     46 * mm),
        ("status",       "Statut",       18 * mm),
        ("created_at",   "Trouvé le",    22 * mm),
    ]
    col_widths = [w for _, _, w in pdf_cols]

    # Données
    def _trim(s: str, n: int) -> str:
        s = (s or "").strip()
        return s if len(s) <= n else s[: n - 1] + "…"

    header = [Paragraph(f"<b>{label}</b>", cell_style)
              for _, label, _ in pdf_cols]
    table_data = [header]
    for r in rows:
        line = []
        for key, _, _ in pdf_cols:
            v = _value_for(r, key)
            if key in ("name", "handle", "website", "emails"):
                v = _trim(v, 70)
            elif key == "platform":
                v = v.capitalize() if v else "—"
            elif key == "created_at":
                v = (v or "")[:10]  # YYYY-MM-DD
            line.append(Paragraph(_escape_xml(v) or "—", cell_style))
        table_data.append(line)

    table = Table(table_data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2D2D44")),
        ("TEXTCOLOR",  (0, 0), (-1, 0), colors.white),
        ("FONTNAME",   (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE",   (0, 0), (-1, 0), 9),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
        ("TOPPADDING",    (0, 0), (-1, 0), 6),
        ("VALIGN",     (0, 0), (-1, -1), "TOP"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1),
         [colors.white, colors.HexColor("#F4F4FA")]),
        ("LINEBELOW",  (0, 0), (-1, -1), 0.25, colors.HexColor("#DDDDE5")),
        ("LEFTPADDING",  (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
    ]))

    now_str = datetime.now().strftime("%d/%m/%Y à %Hh%M")
    flow = [
        Paragraph(f"{title}", title_style),
        Paragraph(f"{len(rows)} prospect(s) — généré le {now_str}", sub_style),
        table,
        Spacer(1, 8),
        Paragraph(
            "Triskell Command · Obelisk — prospection créateurs",
            ParagraphStyle("Foot", parent=styles["Normal"],
                            fontSize=7, textColor=colors.HexColor("#AAA"))
        ),
    ]
    doc.build(flow)
    return buf.getvalue()


def _escape_xml(s: str) -> str:
    """Échappe < > & pour reportlab Paragraph."""
    if not s:
        return ""
    return (s.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;"))
